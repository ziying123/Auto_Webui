#!/usr/bin/env python
# -*- coding: UTF-8 -*-


from openpyxl import load_workbook
from config.conf import DATA_DIR


class DoExcel:
    def __init__(self, FilePath, SheetName):
        self.FilePath = FilePath
        self.SheetName = SheetName

    def get_data(self):
        # # 方法一：
        # test_data = []
        # # print(file_path)
        # wb = load_workbook(self.FilePath)
        # sh = wb[self.SheetName]
        # # print(sh)
        # for r in range(2, sh.max_row + 1):
        #     sub_test_data = {
        #         'CaseID': sh.cell(r, 1).value,
        #         'Title': sh.cell(r, 2).value,
        #         'Data': eval(sh.cell(r, 3).value),
        #         'check': eval(sh.cell(r, 4).value)}
        #     test_data.append(sub_test_data)
        # return test_data

        # 方法二：
        wb = load_workbook(self.FilePath)
        sh = wb[self.SheetName]
        fist_row = []
        test_data = []
        # print(list(sh.rows))
        for item in list(sh.rows)[0]:  # 遍历第1行当中每一列
            fist_row.append(item.value)

        for item in list(sh.rows)[1:]:
            values = []
            for val in item:
                values.append(val.value)
            res = dict(zip(fist_row, values))
            # res["Data"] = eval(res["Data"])
            # res["check"] = eval(res["check"])
            test_data.append(res)
        return test_data

    def close_file(self):
        wb = load_workbook(self.FilePath)
        wb.close()


if __name__ == '__main__':
    # file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")
    do_excel = DoExcel(DATA_DIR, "register")
    res_test_data = do_excel.get_data()
    do_excel.close_file()
    for case in res_test_data:
        print(case)
